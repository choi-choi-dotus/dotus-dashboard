import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import openpyxl
from itertools import combinations
from datetime import date, timedelta
import io

# ── 페이지 설정 ──────────────────────────────────────────
st.set_page_config(
    page_title="Dotus 매출 대시보드",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── 색상 ─────────────────────────────────────────────────
BG      = "#0d1b2a"
CARD    = "#132337"
CARD2   = "#1a2f45"
BORDER  = "#1e3a5f"
CYAN    = "#00BCD4"
CYAN2   = "#4DD0E1"
GOLD    = "#FFC107"
TEXT    = "#E0F7FA"
TEXT2   = "#90A4AE"

CHART_COLORS = [CYAN, "#FFC107", "#FF6B6B", "#AB47BC", "#66BB6A", "#FF9800"]
PIE_COLORS   = [CYAN, CYAN2, "#0097A7", "#00838F", "#006064",
                "#80DEEA", "#B2EBF2", "#E0F7FA", "#4DD0E1", "#26C6DA"]

def nl(**kwargs):
    base = dict(plot_bgcolor=CARD, paper_bgcolor=CARD,
                font=dict(color=TEXT2, size=12), margin=dict(t=40,b=20,l=10,r=10))
    base.update(kwargs)
    return base

AXIS = dict(gridcolor=CARD2, linecolor=BORDER, tickcolor=BORDER, tickfont=dict(color=TEXT2))

# ── CSS ──────────────────────────────────────────────────
st.markdown(f"""
<style>
.stApp {{ background-color: {BG}; }}
[data-testid="stAppViewContainer"] > .main {{ background-color: {BG}; }}
[data-testid="stSidebar"] {{
    background-color: #0a1628;
    border-right: 1px solid {BORDER};
}}
[data-testid="stSidebar"] * {{ color: {TEXT2} !important; }}
[data-testid="stSidebar"] h2 {{ color: {TEXT} !important; font-size:1rem !important; font-weight:600 !important; }}
[data-testid="stSidebar"] .stRadio div[role="radiogroup"] label {{
    padding: 10px 14px !important;
    border-radius: 8px !important;
    font-size: 0.92rem !important;
    font-weight: 500 !important;
    cursor: pointer !important;
    display: flex !important;
    align-items: center !important;
    color: {TEXT2} !important;
    margin-bottom: 4px !important;
}}
p, span, label, div {{ color: {TEXT2}; }}
h1, h2, h3, h4 {{ color: {TEXT} !important; }}
hr {{ border-color: {BORDER} !important; }}
[data-testid="stMetric"] {{
    background: {CARD};
    border: 1px solid {BORDER};
    border-top: 3px solid {CYAN};
    border-radius: 8px;
    padding: 20px 24px;
}}
[data-testid="stMetricValue"] {{ font-size:1.8rem !important; font-weight:700 !important; color:{CYAN} !important; }}
[data-testid="stMetricLabel"] {{ font-size:0.78rem !important; color:{TEXT2} !important; text-transform:uppercase; letter-spacing:0.05em; }}
span[data-baseweb="tag"] {{ background-color: #0e3a50 !important; color:{CYAN} !important; }}
span[data-baseweb="tag"] span {{ color:{CYAN} !important; }}
.stRadio label {{ color:{TEXT2} !important; font-size:0.88rem !important; }}
.stButton button {{
    background-color: {CARD} !important; color:{CYAN} !important;
    border: 1px solid {BORDER} !important; border-radius:4px !important;
}}
.stButton button:hover {{ border-color:{CYAN} !important; background-color:#0e3a50 !important; }}
[data-testid="stDataFrame"] {{ border:1px solid {BORDER}; border-radius:8px; }}
.stTextInput input {{
    background-color: {CARD} !important;
    border: 1px solid {BORDER} !important;
    color: {TEXT} !important;
    border-radius: 6px !important;
}}
.stSelectbox > div > div {{
    background-color: {CARD} !important;
    border: 1px solid {BORDER} !important;
    color: {TEXT} !important;
}}
</style>
""", unsafe_allow_html=True)

# ── 비밀번호 ──────────────────────────────────────────────
PASSWORD = "dotus2026"

def check_password():
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False
    if st.session_state.authenticated:
        return True
    st.markdown(f"""
    <div style='display:flex;justify-content:center;margin-top:80px;'>
        <div style='text-align:center;'>
            <h1 style='font-size:2.2rem;font-weight:700;color:{CYAN};margin-bottom:4px;'>📊 Dotus</h1>
            <p style='color:{TEXT2};margin-bottom:32px;'>매출 대시보드</p>
        </div>
    </div>""", unsafe_allow_html=True)
    c1, c2, c3 = st.columns([1,1.2,1])
    with c2:
        pw = st.text_input("비밀번호", type="password", placeholder="비밀번호를 입력하세요")
        if st.button("로그인", use_container_width=True, type="primary"):
            if pw == PASSWORD:
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("비밀번호가 틀렸습니다.")
    return False

if not check_password():
    st.stop()

# ── 데이터 로드 ───────────────────────────────────────────
@st.cache_data
def load_data():
    wb = openpyxl.load_workbook("data.xlsx", read_only=True, data_only=True)

    ws1 = wb["시트1"]
    rows1 = list(ws1.iter_rows(values_only=True))
    headers1 = rows1[0][:16]
    data1 = [r[:16] for r in rows1[1:] if r[0] is not None]
    df = pd.DataFrame(data1, columns=headers1)
    df["date"]       = pd.to_datetime(df["date"])
    df["year_month"] = df["date"].dt.to_period("M").astype(str)
    df["net_sales"]  = df["sales_amount"] - df["return_sales_amount"]
    df["net_qty"]    = df["quantity"] - df["return_quantity"]

    ws2 = wb["시트2"]
    rows2 = list(ws2.iter_rows(values_only=True))
    headers2 = rows2[0][:13]
    data2 = [r[:13] for r in rows2[1:] if r[0] is not None]
    df_stock = pd.DataFrame(data2, columns=headers2)
    df_stock["date"]       = pd.to_datetime(df_stock["date"])
    df_stock["year_month"] = df_stock["date"].dt.to_period("M").astype(str)

    ws3 = wb["품목마스터"]
    rows3 = list(ws3.iter_rows(values_only=True))
    headers3 = rows3[0][:8]
    data3 = [r[:8] for r in rows3[1:] if r[0] is not None]
    df_master = pd.DataFrame(data3, columns=headers3)

    return df, df_stock, df_master

df_sales, df_stock, df_master = load_data()
HAS_SMALL = "small_category" in df_sales.columns

# ── 전체 선택 멀티셀렉트 헬퍼 ────────────────────────────
def multiselect_with_all(label, options, key, **kwargs):
    ALL = "전체"
    choices = [ALL] + list(options)
    selected = st.multiselect(label, options=choices, default=[ALL], key=key, **kwargs)
    if not selected or ALL in selected:
        return list(options)
    return selected

# ── 챗봇 답변 함수 ───────────────────────────────────────
def answer_question(q, df, today):
    import re
    q_orig  = q.strip()
    q_lower = q_orig.lower()

    # ── 날짜 추출 ─────────────────────────────────────────
    date_filter  = None
    period_label = None

    # N월 M일부터 N월 K일까지 / N월 M일~N월 K일
    m = re.search(r'(\d+)월\s*(\d+)일\s*(?:부터|~|∼|-)\s*(\d+)월\s*(\d+)일', q_orig)
    if m:
        try:
            d1 = date(today.year, int(m.group(1)), int(m.group(2)))
            d2 = date(today.year, int(m.group(3)), int(m.group(4)))
            if d1 > d2: d1, d2 = d2, d1
            date_filter  = (d1, d2)
            period_label = f"{m.group(1)}월 {m.group(2)}일 ~ {m.group(3)}월 {m.group(4)}일"
        except: pass

    # N월 M일부터 K일까지 (같은 달)
    if not date_filter:
        m = re.search(r'(\d+)월\s*(\d+)일\s*(?:부터|~|∼|-)\s*(\d+)일', q_orig)
        if m:
            try:
                mo = int(m.group(1))
                d1 = date(today.year, mo, int(m.group(2)))
                d2 = date(today.year, mo, int(m.group(3)))
                if d1 > d2: d1, d2 = d2, d1
                date_filter  = (d1, d2)
                period_label = f"{mo}월 {m.group(2)}일 ~ {m.group(3)}일"
            except: pass

    # N월부터 M월까지 / N월~M월
    if not date_filter:
        m = re.search(r'(\d+)월\s*(?:부터|~|∼|-)\s*(\d+)월', q_orig)
        if m:
            try:
                mo1, mo2 = int(m.group(1)), int(m.group(2))
                if mo1 > mo2: mo1, mo2 = mo2, mo1
                d1 = date(today.year, mo1, 1)
                d2 = date(today.year, mo2+1, 1) - timedelta(days=1) if mo2 < 12 else date(today.year, 12, 31)
                date_filter  = (d1, d2)
                period_label = f"{mo1}월 ~ {mo2}월"
            except: pass

    # N월 M일
    if not date_filter:
        m = re.search(r'(\d+)월\s*(\d+)일', q_orig)
        if m:
            try:
                d = date(today.year, int(m.group(1)), int(m.group(2)))
                date_filter  = (d, d)
                period_label = f"{today.year}년 {m.group(1)}월 {m.group(2)}일"
            except: pass

    # N월 (단독)
    if not date_filter:
        m = re.search(r'(\d+)월(?!\s*\d+일)', q_orig)
        if m:
            mo = int(m.group(1))
            try:
                first = date(today.year, mo, 1)
                last  = date(today.year, mo+1, 1) - timedelta(days=1) if mo < 12 else date(today.year, 12, 31)
                date_filter  = (first, last)
                period_label = f"{today.year}년 {mo}월"
            except: pass

    if not date_filter and '오늘' in q_orig:
        date_filter = (today, today); period_label = "오늘"

    if not date_filter and '이번달' in q_orig:
        date_filter = (today.replace(day=1), today); period_label = "이번달"

    if not date_filter and '지난달' in q_orig:
        ft = today.replace(day=1)
        lp = ft - timedelta(days=1)
        date_filter = (lp.replace(day=1), lp); period_label = "지난달"

    if not date_filter:
        m = re.search(r'최근\s*(\d+)일', q_orig)
        if m:
            n = int(m.group(1))
            date_filter = (today - timedelta(days=n), today); period_label = f"최근 {n}일"

    # ── 데이터 필터 ────────────────────────────────────────
    filtered = df.copy()
    if date_filter:
        filtered = filtered[
            (filtered["date"].dt.date >= date_filter[0]) &
            (filtered["date"].dt.date <= date_filter[1])
        ]
    period_label = period_label or "전체 기간"

    # 채널 필터
    sel_client = None
    for c in sorted(df["client"].unique(), key=len, reverse=True):
        if c in q_orig:
            sel_client = c
            filtered = filtered[filtered["client"] == c]
            break

    # 상품 필터
    sel_product = None
    for p in sorted(df["product_name"].dropna().unique(), key=len, reverse=True):
        if p in q_orig:
            sel_product = p
            filtered = filtered[filtered["product_name"] == p]
            break

    ctx = ""
    if sel_client:  ctx += f"  [{sel_client}]"
    if sel_product: ctx += f"  [{sel_product}]"
    header = f"**{period_label}{ctx}**"

    if filtered.empty:
        return f"{header}\n\n해당 기간/조건에 데이터가 없습니다."

    # ── TOP N 숫자 추출 ───────────────────────────────────
    n_top = 5
    m = re.search(r'(?:top|상위|최고)\s*(\d+)|(\d+)(?:개|위)\s*(?:순위|랭킹)', q_lower)
    if m:
        for g in m.groups():
            if g: n_top = int(g); break

    # ── 최대/최소값 질문 ─────────────────────────────────
    is_max = any(k in q_lower for k in ['가장 높','가장 많','최고','최대','피크','제일 높','제일 많'])
    is_min = any(k in q_lower for k in ['가장 낮','가장 적','최저','최소','제일 낮','제일 적'])
    by_qty = any(k in q_lower for k in ['수량','개수','출고'])

    if is_max or is_min:
        col = "net_qty" if by_qty else "net_sales"
        lbl = "판매수량" if by_qty else "결제금액"
        asc = is_min

        # 날짜 기준
        if any(k in q_lower for k in ['날짜','날','일자','하루']):
            agg = filtered.groupby("date")[col].sum().reset_index()
            row = agg.loc[agg[col].idxmin() if asc else agg[col].idxmax()]
            direction = "최저" if asc else "최고"
            val = f"{row[col]:,}개" if by_qty else f"₩{row[col]/1e8:.2f}억"
            return (f"📅 {header} {lbl} {direction} 날짜\n\n"
                    f"**{row['date'].strftime('%Y년 %m월 %d일')}**\n\n{val}")

        # 월 기준
        if any(k in q_lower for k in ['월','달']):
            agg = filtered.groupby("year_month")[col].sum().reset_index()
            row = agg.loc[agg[col].idxmin() if asc else agg[col].idxmax()]
            direction = "최저" if asc else "최고"
            val = f"{row[col]:,}개" if by_qty else f"₩{row[col]/1e8:.2f}억"
            return (f"📅 {header} {lbl} {direction} 월\n\n"
                    f"**{row['year_month']}**\n\n{val}")

        # 상품 기준
        if any(k in q_lower for k in ['상품','품목','제품']):
            agg = filtered.groupby("product_name")[col].sum().reset_index()
            row = agg.loc[agg[col].idxmin() if asc else agg[col].idxmax()]
            direction = "최저" if asc else "최고"
            val = f"{row[col]:,}개" if by_qty else f"₩{row[col]/1e8:.2f}억"
            return (f"🏆 {header} {lbl} {direction} 상품\n\n"
                    f"**{row['product_name']}**\n\n{val}")

        # 채널 기준
        if any(k in q_lower for k in ['채널','채날']):
            agg = filtered.groupby("client")[col].sum().reset_index()
            row = agg.loc[agg[col].idxmin() if asc else agg[col].idxmax()]
            direction = "최저" if asc else "최고"
            val = f"{row[col]:,}개" if by_qty else f"₩{row[col]/1e8:.2f}억"
            return (f"🏆 {header} {lbl} {direction} 채널\n\n"
                    f"**{row['client']}**\n\n{val}")

        # 기본 = 날짜 기준
        agg = filtered.groupby("date")[col].sum().reset_index()
        row = agg.loc[agg[col].idxmin() if asc else agg[col].idxmax()]
        direction = "최저" if asc else "최고"
        val = f"{row[col]:,}개" if by_qty else f"₩{row[col]/1e8:.2f}억"
        return (f"📅 {header} {lbl} {direction} 날짜\n\n"
                f"**{row['date'].strftime('%Y년 %m월 %d일')}**\n\n{val}")

    # ── 채널별 ────────────────────────────────────────────
    if '채널별' in q_orig or '채널 별' in q_orig:
        agg = filtered.groupby("client").agg(
            결제금액=("net_sales","sum"), 수량=("net_qty","sum")
        ).sort_values("결제금액", ascending=False).reset_index()
        lines = [f"📊 {header} 채널별 결제금액\n"]
        for i, row in agg.iterrows():
            lines.append(f"{i+1}. {row['client']}: ₩{row['결제금액']/1e8:.2f}억 ({row['수량']:,}개)")
        return "\n".join(lines)

    # ── 상품 순위 ─────────────────────────────────────────
    rank_kw = ['순위','top','랭킹','많이 팔','잘 나가','상위','베스트']
    if any(k in q_lower for k in rank_kw) and not sel_product:
        by_qty = any(k in q_lower for k in ['수량','개수'])
        col    = "net_qty" if by_qty else "net_sales"
        lbl    = "판매수량" if by_qty else "결제금액"
        agg    = filtered.groupby("product_name")[col].sum().sort_values(ascending=False).head(n_top).reset_index()
        lines  = [f"🏆 {header} 상품 {lbl} TOP {n_top}\n"]
        for i, row in agg.iterrows():
            val = f"{row[col]:,}개" if by_qty else f"₩{row[col]/1e4:.0f}만"
            lines.append(f"{i+1}. {row['product_name']}: {val}")
        return "\n".join(lines)

    # ── 수량 ──────────────────────────────────────────────
    if any(k in q_lower for k in ['수량','몇개','몇 개','판매량','출고']):
        total = filtered["net_qty"].sum()
        return f"📦 {header} 판매수량\n\n**{total:,}개**"

    # ── 주문건수 ──────────────────────────────────────────
    if any(k in q_lower for k in ['주문','건수','몇건','몇 건']):
        total = filtered["order_no_1"].nunique()
        return f"🧾 {header} 주문건수\n\n**{total:,}건**"

    # ── 기본 = 결제금액 ───────────────────────────────────
    sales  = filtered["net_sales"].sum()
    qty    = filtered["net_qty"].sum()
    orders = filtered["order_no_1"].nunique()
    return (f"💰 {header} 결제금액\n\n"
            f"**₩{sales/1e8:.2f}억**\n\n"
            f"판매수량 {qty:,}개  ·  주문건수 {orders:,}건")

# ── 사이드바 메뉴 ─────────────────────────────────────────
with st.sidebar:
    st.markdown("## 📊 Dotus")
    st.markdown("---")
    page = st.radio(
        "메뉴",
        ["📈 매출 대시보드", "🔍 상세 데이터 조회", "🧾 주문 상세분석", "📦 재고소진일정", "💬 챗봇"],
        label_visibility="collapsed"
    )
    st.markdown("---")
    if st.button("🚪 로그아웃", use_container_width=True):
        st.session_state.authenticated = False
        st.rerun()

# ════════════════════════════════════════════════════════
# PAGE 1 : 매출 대시보드
# ════════════════════════════════════════════════════════
if page == "📈 매출 대시보드":

    st.markdown("# 📈 매출 대시보드")
    st.markdown("---")

    with st.expander("🔧 조회 조건", expanded=True):
        r1c1, r1c2, r1c3 = st.columns(3)
        with r1c1:
            months = sorted(df_sales["year_month"].unique())
            month_start = st.selectbox("시작 월", months, index=0)
            month_end   = st.selectbox("종료 월", months, index=len(months)-1)
        with r1c2:
            all_clients = sorted(df_sales["client"].unique())
            selected_clients = multiselect_with_all("채널", all_clients, key="p1_ch")
        with r1c3:
            all_large = sorted(df_sales["large_category"].dropna().unique())
            selected_large = multiselect_with_all("대분류", all_large, key="p1_lg")
            mid_pool = df_sales[df_sales["large_category"].isin(selected_large)]["medium_category"].dropna().unique()
            selected_medium = multiselect_with_all("중분류", sorted(mid_pool), key="p1_md")
            selected_small = None
            if HAS_SMALL:
                small_pool = df_sales[
                    df_sales["large_category"].isin(selected_large) &
                    df_sales["medium_category"].isin(selected_medium)
                ]["small_category"].dropna().unique()
                selected_small = multiselect_with_all("소분류", sorted(small_pool), key="p1_sm")

    filtered = df_sales[
        (df_sales["year_month"] >= month_start) &
        (df_sales["year_month"] <= month_end) &
        (df_sales["client"].isin(selected_clients)) &
        (df_sales["large_category"].isin(selected_large)) &
        (df_sales["medium_category"].isin(selected_medium))
    ]
    if HAS_SMALL and selected_small:
        filtered = filtered[filtered["small_category"].isin(selected_small)]

    prev_months = sorted(df_sales["year_month"].unique())
    if month_start in prev_months and prev_months.index(month_start) > 0:
        ps = prev_months[max(0, prev_months.index(month_start)-1)]
        pe = prev_months[max(0, prev_months.index(month_end)-1)]
        prev_f = df_sales[
            (df_sales["year_month"] >= ps) & (df_sales["year_month"] <= pe) &
            (df_sales["client"].isin(selected_clients)) &
            (df_sales["large_category"].isin(selected_large))
        ]
    else:
        prev_f = pd.DataFrame()

    total_sales  = filtered["net_sales"].sum()
    total_qty    = filtered["net_qty"].sum()
    total_orders = filtered["order_no_1"].nunique()
    avg_order    = total_sales / total_orders if total_orders > 0 else 0

    if not prev_f.empty:
        pss = prev_f["net_sales"].sum(); pq = prev_f["net_qty"].sum(); po = prev_f["order_no_1"].nunique()
        sales_d  = f"{((total_sales -pss)/pss*100):+.1f}%" if pss else "N/A"
        qty_d    = f"{((total_qty  -pq )/pq *100):+.1f}%" if pq  else "N/A"
        orders_d = f"{((total_orders-po )/po *100):+.1f}%" if po  else "N/A"
    else:
        sales_d = qty_d = orders_d = None

    k1,k2,k3,k4 = st.columns(4)
    with k1: st.metric("결제금액",      f"₩{total_sales/1e8:.1f}억",  delta=sales_d)
    with k2: st.metric("판매 수량",     f"{total_qty:,}개",            delta=qty_d)
    with k3: st.metric("주문 건수",     f"{total_orders:,}건",         delta=orders_d)
    with k4: st.metric("평균 주문금액", f"₩{avg_order/1e4:.0f}만")
    st.markdown("---")

    st.markdown("### 월별 결제금액 추이")
    sales_metric = st.radio("기준", ["금액 (억원)", "수량 (개)"], horizontal=True, key="sm1")
    monthly = filtered.groupby("year_month").agg(결제금액=("net_sales","sum"), 수량=("net_qty","sum")).reset_index()
    use_m = (sales_metric == "금액 (억원)")
    y_v   = monthly["결제금액"]/1e8 if use_m else monthly["수량"]
    y_fmt = [f"₩{v:.1f}억" for v in y_v] if use_m else [f"{v:,}개" for v in y_v]
    monthly["MoM"] = y_v.pct_change() * 100

    fig_sales = make_subplots(specs=[[{"secondary_y": True}]])
    fig_sales.add_trace(go.Scatter(
        x=monthly["year_month"], y=y_v, name="결제금액",
        mode="lines+markers", line=dict(color=CYAN, width=3, shape="spline"),
        marker=dict(size=7, color=CYAN), fill="tozeroy", fillcolor="rgba(0,188,212,0.12)",
        text=y_fmt, hovertemplate="%{text}<extra></extra>"
    ), secondary_y=False)
    fig_sales.add_trace(go.Scatter(
        x=monthly["year_month"], y=monthly["MoM"], name="MoM(%)",
        mode="lines+markers", line=dict(color=GOLD, width=2, dash="dot", shape="spline"),
        marker=dict(size=5, color=GOLD), hovertemplate="%{y:.1f}%<extra></extra>"
    ), secondary_y=True)
    fig_sales.update_layout(height=340, **nl(legend=dict(orientation="h", yanchor="bottom", y=1.02, font=dict(color=TEXT2))))
    fig_sales.update_xaxes(type="category", **AXIS)
    fig_sales.update_yaxes(title_text="억원" if use_m else "수량", secondary_y=False, **AXIS)
    fig_sales.update_yaxes(title_text="MoM (%)", secondary_y=True, **AXIS)
    st.plotly_chart(fig_sales, use_container_width=True)

    cl, cr = st.columns(2)
    with cl:
        st.markdown("### 채널별 결제금액")
        ch_df = filtered.groupby("client").agg(결제금액=("net_sales","sum")).sort_values("결제금액", ascending=True).reset_index()
        fig_ch = go.Figure(go.Bar(
            x=ch_df["결제금액"]/1e8, y=ch_df["client"], orientation="h",
            marker=dict(color=ch_df["결제금액"], colorscale=[[0,CARD2],[1,CYAN]], showscale=False),
            text=[f"₩{v/1e8:.1f}억" for v in ch_df["결제금액"]],
            textposition="outside", textfont=dict(color=TEXT2)
        ))
        fig_ch.update_layout(height=380, **nl(margin=dict(t=20,b=20,l=10,r=80)))
        fig_ch.update_xaxes(title_text="억원", **AXIS)
        fig_ch.update_yaxes(**AXIS)
        st.plotly_chart(fig_ch, use_container_width=True)

    with cr:
        st.markdown("### 상품 중분류별 결제금액")
        mid_df = filtered.groupby("medium_category").agg(결제금액=("net_sales","sum")).sort_values("결제금액", ascending=False).head(10).reset_index()
        fig_mid = go.Figure(go.Pie(
            values=mid_df["결제금액"], labels=mid_df["medium_category"],
            marker=dict(colors=PIE_COLORS, line=dict(color=BG, width=2)),
            textposition="inside", textinfo="percent+label",
            textfont=dict(size=11, color=BG), hole=0.4
        ))
        fig_mid.update_layout(height=380, plot_bgcolor=CARD, paper_bgcolor=CARD,
                              font=dict(color=TEXT2), margin=dict(t=20,b=20,l=10,r=10), showlegend=False)
        st.plotly_chart(fig_mid, use_container_width=True)

    st.markdown("### 채널별 월별 결제금액 흐름")
    ch_monthly = filtered.groupby(["year_month","client"])["net_sales"].sum().reset_index()
    top6 = ch_df.sort_values("결제금액", ascending=False).head(6)["client"].tolist()
    fig_multi = go.Figure()
    for i, client in enumerate(top6):
        df_c = ch_monthly[ch_monthly["client"]==client]
        color = CHART_COLORS[i % len(CHART_COLORS)]
        r,g,b = int(color[1:3],16), int(color[3:5],16), int(color[5:7],16)
        fig_multi.add_trace(go.Scatter(
            x=df_c["year_month"], y=df_c["net_sales"]/1e8, name=client,
            mode="lines+markers", line=dict(color=color, width=2, shape="spline"),
            marker=dict(size=5, color=color), fill="tozeroy", fillcolor=f"rgba({r},{g},{b},0.06)",
            hovertemplate=f"{client}: ₩%{{y:.1f}}억<extra></extra>"
        ))
    fig_multi.update_layout(height=340, **nl(legend=dict(orientation="h", yanchor="bottom", y=1.02, font=dict(color=TEXT2))))
    fig_multi.update_xaxes(type="category", **AXIS)
    fig_multi.update_yaxes(title_text="억원", **AXIS)
    st.plotly_chart(fig_multi, use_container_width=True)

    st.markdown("### 채널 × 월별 결제금액 히트맵")
    piv = filtered.groupby(["client","year_month"])["net_sales"].sum().reset_index()
    piv_t = piv.pivot(index="client", columns="year_month", values="net_sales").fillna(0)
    fig_heat = go.Figure(go.Heatmap(
        z=piv_t.values/1e8, x=piv_t.columns.tolist(), y=piv_t.index.tolist(),
        colorscale=[[0,CARD2],[0.5,"#00838F"],[1,CYAN]],
        text=[[f"₩{v:.1f}억" for v in row] for row in piv_t.values/1e8],
        texttemplate="%{text}", textfont=dict(color=TEXT, size=11),
        colorbar=dict(title=dict(text="억원", font=dict(color=TEXT2)), tickfont=dict(color=TEXT2))
    ))
    fig_heat.update_layout(height=340, plot_bgcolor=CARD, paper_bgcolor=CARD,
                           font=dict(color=TEXT2), margin=dict(t=20,b=20,l=10,r=10),
                           xaxis=dict(tickfont=dict(color=TEXT2)), yaxis=dict(tickfont=dict(color=TEXT2)))
    st.plotly_chart(fig_heat, use_container_width=True)

    st.markdown("---")

    sv_opt = st.radio("매출 표시 기준", ["금액 (억원)", "수량 (개)"], horizontal=True, key="sv2")
    lc, rc = st.columns(2)
    CHART_H = 300

    with lc:
        st.markdown("### 월별 결제금액")
        sm2 = filtered.groupby("year_month").agg(결제금액=("net_sales","sum"), 수량=("net_qty","sum")).reset_index()
        use_m2 = (sv_opt == "금액 (억원)")
        sv_y  = sm2["결제금액"]/1e8 if use_m2 else sm2["수량"]
        sv_fmt= [f"₩{v:.1f}억" for v in sv_y] if use_m2 else [f"{v:,}개" for v in sv_y]
        fig_sv = go.Figure(go.Scatter(
            x=sm2["year_month"], y=sv_y, mode="lines+markers",
            line=dict(color=CYAN, width=2.5, shape="spline"), marker=dict(size=7, color=CYAN),
            fill="tozeroy", fillcolor="rgba(0,188,212,0.10)",
            text=sv_fmt, hovertemplate="%{text}<extra></extra>"
        ))
        fig_sv.update_layout(height=CHART_H, **nl())
        fig_sv.update_xaxes(type="category", **AXIS)
        fig_sv.update_yaxes(title_text="억원" if use_m2 else "수량", **AXIS)
        st.plotly_chart(fig_sv, use_container_width=True)

    with rc:
        st.markdown("### 월별 입고 수량")
        stk = df_stock[
            (df_stock["year_month"] >= month_start) &
            (df_stock["year_month"] <= month_end)
        ].groupby("year_month")["quantity"].sum().reset_index()
        stk.columns = ["year_month","입고수량"]
        fig_stk = go.Figure(go.Scatter(
            x=stk["year_month"], y=stk["입고수량"], mode="lines+markers",
            line=dict(color=CYAN2, width=2.5, shape="spline"), marker=dict(size=7, color=CYAN2),
            fill="tozeroy", fillcolor="rgba(77,208,225,0.10)",
            hovertemplate="입고: %{y:,}개<extra></extra>"
        ))
        fig_stk.update_layout(height=CHART_H, **nl())
        fig_stk.update_xaxes(type="category", **AXIS)
        fig_stk.update_yaxes(title_text="수량 (개)", tickformat=",", **AXIS)
        st.plotly_chart(fig_stk, use_container_width=True)

# ════════════════════════════════════════════════════════
# PAGE 2 : 상세 데이터 조회
# ════════════════════════════════════════════════════════
elif page == "🔍 상세 데이터 조회":

    st.markdown("# 🔍 상세 데이터 조회")
    st.markdown("---")

    with st.expander("🔧 조회 조건", expanded=True):
        d1, d2, d3 = st.columns([1,1,2])
        with d1:
            date_from = st.date_input("시작일", value=df_sales["date"].min().date(),
                                      min_value=df_sales["date"].min().date(),
                                      max_value=df_sales["date"].max().date())
        with d2:
            date_to = st.date_input("종료일", value=df_sales["date"].max().date(),
                                    min_value=df_sales["date"].min().date(),
                                    max_value=df_sales["date"].max().date())
        with d3:
            det_clients = multiselect_with_all("채널", sorted(df_sales["client"].unique()), key="det_ch")
        keyword = st.text_input("상품명 검색", placeholder="예: 오딧  →  오딧 캐리어, 오딧백 등 모두 검색")

    detail_df = df_sales[
        (df_sales["date"].dt.date >= date_from) &
        (df_sales["date"].dt.date <= date_to) &
        (df_sales["client"].isin(det_clients))
    ].copy()
    if keyword.strip():
        for kw in keyword.strip().split():
            detail_df = detail_df[detail_df["product_name"].str.contains(kw, case=False, na=False)]

    # 조회 일수
    days = (date_to - date_from).days + 1
    total_sales_sum = detail_df["net_sales"].sum()
    total_qty_sum   = detail_df["net_qty"].sum()
    avg_qty_day     = total_qty_sum / days
    avg_sales_day   = total_sales_sum / days

    c1,c2,c3,c4,c5 = st.columns(5)
    with c1: st.metric("조회 건수",      f"{len(detail_df):,}건")
    with c2: st.metric("결제금액 합계",  f"₩{total_sales_sum/1e8:.2f}억")
    with c3: st.metric("판매 수량",      f"{total_qty_sum:,}개")
    with c4: st.metric("일평균 출고량",  f"{avg_qty_day:.1f}개")
    with c5: st.metric("일평균 결제금액", f"₩{avg_sales_day/1e4:.0f}만")

    st.markdown("---")

    st.markdown("### 조회 기간 결제금액 흐름")
    spark = detail_df.groupby("year_month")["net_sales"].sum().reset_index().sort_values("year_month")
    if not spark.empty:
        fig_sp = go.Figure(go.Scatter(
            x=spark["year_month"], y=spark["net_sales"]/1e8,
            mode="lines+markers", line=dict(color=CYAN, width=3, shape="spline"),
            marker=dict(size=7, color=CYAN, line=dict(color=BG, width=2)),
            fill="tozeroy", fillcolor="rgba(0,188,212,0.12)",
            text=[f"₩{v:.2f}억" for v in spark["net_sales"]/1e8],
            hovertemplate="%{x}: %{text}<extra></extra>"
        ))
        fig_sp.update_layout(height=220, **nl(margin=dict(t=20,b=20,l=10,r=10)))
        fig_sp.update_xaxes(type="category", **AXIS)
        fig_sp.update_yaxes(title_text="억원", **AXIS)
        st.plotly_chart(fig_sp, use_container_width=True)

    st.markdown("---")

    st.markdown("### 상품별 기간 합산")
    sort_by = st.radio("정렬 기준", ["결제금액", "판매수량"], horizontal=True)

    agg = detail_df.groupby("product_name").agg(
        결제금액=("net_sales","sum"), 판매수량=("net_qty","sum"),
        주문건수=("order_no_1","nunique"), 환불수량=("return_quantity","sum"),
        환불금액=("return_sales_amount","sum")
    ).reset_index().rename(columns={"product_name":"상품명"})

    agg["평균출고량"]   = (agg["판매수량"] / days).round(1)
    agg["평균결제금액"] = (agg["결제금액"] / days).round(0).astype(int)

    agg = agg.sort_values(sort_by, ascending=False).reset_index(drop=True)
    agg.index += 1

    disp = agg.copy()
    disp["결제금액"]   = disp["결제금액"].apply(lambda x: f"₩{x:,}")
    disp["환불금액"]   = disp["환불금액"].apply(lambda x: f"₩{x:,}")
    disp["평균결제금액"] = disp["평균결제금액"].apply(lambda x: f"₩{x:,}")
    disp = disp[["상품명","결제금액","판매수량","평균출고량","평균결제금액","주문건수","환불수량","환불금액"]]
    st.dataframe(disp, use_container_width=True, hide_index=False, height=480)

# ════════════════════════════════════════════════════════
# PAGE 3 : 주문 상세분석
# ════════════════════════════════════════════════════════
elif page == "🧾 주문 상세분석":

    st.markdown("# 🧾 주문 상세분석")
    st.markdown("---")

    with st.expander("🔧 조회 조건", expanded=True):
        o1, o2, o3, o4 = st.columns([1,1,2,2])
        with o1:
            o_from = st.date_input("시작일", value=df_sales["date"].min().date(),
                                   min_value=df_sales["date"].min().date(),
                                   max_value=df_sales["date"].max().date(), key="o_from")
        with o2:
            o_to = st.date_input("종료일", value=df_sales["date"].max().date(),
                                 min_value=df_sales["date"].min().date(),
                                 max_value=df_sales["date"].max().date(), key="o_to")
        with o3:
            o_clients = multiselect_with_all("채널", sorted(df_sales["client"].unique()), key="o_ch")
        with o4:
            order_kw = st.text_input("주문번호 검색", placeholder="주문번호 일부만 입력해도 검색 가능")

    ord_df = df_sales[
        (df_sales["date"].dt.date >= o_from) &
        (df_sales["date"].dt.date <= o_to) &
        (df_sales["client"].isin(o_clients))
    ].copy()
    if order_kw.strip():
        ord_df = ord_df[ord_df["order_no_1"].astype(str).str.contains(order_kw.strip(), case=False, na=False)]

    ord_df["order_no_1_str"] = ord_df["order_no_1"].astype(str).str.strip()
    unknown_mask = ord_df["order_no_1_str"].isin(["-", "", "nan", "None"])
    ord_known = ord_df[~unknown_mask].copy()
    ord_unknown = ord_df[unknown_mask].copy()

    cnt_per_order = ord_known.groupby("order_no_1_str").size().reset_index(name="item_count")
    single_orders = cnt_per_order[cnt_per_order["item_count"] == 1]["order_no_1_str"]
    multi_orders  = cnt_per_order[cnt_per_order["item_count"] >= 2]["order_no_1_str"]

    n_single  = len(single_orders)
    n_multi   = len(multi_orders)
    n_unknown = ord_unknown["order_no_1_str"].nunique()
    n_total   = n_single + n_multi + n_unknown

    k1,k2,k3,k4 = st.columns(4)
    with k1: st.metric("전체 주문", f"{n_total:,}건")
    with k2: st.metric("단수구매",   f"{n_single:,}건  ({n_single/n_total*100:.1f}%)" if n_total else "0건")
    with k3: st.metric("복수구매",   f"{n_multi:,}건  ({n_multi/n_total*100:.1f}%)"  if n_total else "0건")
    with k4: st.metric("미확인",   f"{n_unknown:,}건")

    st.markdown("---")

    cl_d, cr_d = st.columns([1, 2])

    with cl_d:
        st.markdown("### 포장 유형 비율")
        fig_donut = go.Figure(go.Pie(
            values=[n_single, n_multi, n_unknown],
            labels=["단수구매", "복수구매", "미확인"],
            marker=dict(colors=[CYAN, GOLD, TEXT2], line=dict(color=BG, width=2)),
            textinfo="percent+label", textfont=dict(size=13, color=BG),
            hole=0.5
        ))
        fig_donut.update_layout(
            height=320, plot_bgcolor=CARD, paper_bgcolor=CARD,
            font=dict(color=TEXT2), margin=dict(t=20,b=20,l=10,r=10),
            showlegend=False,
            annotations=[dict(text=f"{n_total:,}건", x=0.5, y=0.5,
                              font=dict(size=16, color=TEXT), showarrow=False)]
        )
        st.plotly_chart(fig_donut, use_container_width=True)

    with cr_d:
        st.markdown("### 월별 단수구매 / 복수구매 추이")
        ord_known2 = ord_known.copy()
        ord_known2["pack_type"] = ord_known2["order_no_1_str"].isin(multi_orders).map({True:"복수구매", False:"단수구매"})
        pack_monthly = ord_known2.groupby(["year_month","pack_type"])["order_no_1_str"].nunique().reset_index()
        pack_monthly.columns = ["year_month","pack_type","건수"]

        fig_pack = go.Figure()
        for pt, color in [("단수구매", CYAN), ("복수구매", GOLD)]:
            df_pt = pack_monthly[pack_monthly["pack_type"]==pt]
            fig_pack.add_trace(go.Scatter(
                x=df_pt["year_month"], y=df_pt["건수"], name=pt,
                mode="lines+markers", line=dict(color=color, width=2.5, shape="spline"),
                marker=dict(size=6, color=color),
                hovertemplate=f"{pt}: %{{y:,}}건<extra></extra>"
            ))
        fig_pack.update_layout(height=320, **nl(legend=dict(orientation="h", yanchor="bottom", y=1.02, font=dict(color=TEXT2))))
        fig_pack.update_xaxes(type="category", **AXIS)
        fig_pack.update_yaxes(tickformat=",", **AXIS)
        st.plotly_chart(fig_pack, use_container_width=True)

    st.markdown("---")

    st.markdown("### 복수구매 상품 조합 분석")
    multi_df = ord_known[ord_known["order_no_1_str"].isin(multi_orders)].copy()

    if multi_df.empty:
        st.info("복수구매 데이터가 없습니다.")
    else:
        all_products = sorted(multi_df["product_name"].dropna().unique().tolist())
        search_kw = st.text_input("상품명 검색", placeholder="예: 오딧 캐리어", key="prod_search")
        filtered_products = [p for p in all_products if search_kw.strip().lower() in p.lower()] if search_kw.strip() else all_products

        if not filtered_products:
            st.warning("검색 결과가 없습니다.")
        else:
            sel_product = st.selectbox("상품 선택", options=filtered_products, key="prod_select")

            # 선택 상품이 포함된 합포 주문 추출
            orders_with = multi_df[multi_df["product_name"] == sel_product]["order_no_1_str"].unique()
            n_orders_with = len(orders_with)

            # 그 주문들에서 선택 상품 제외한 나머지 상품 집계
            paired_df = multi_df[
                (multi_df["order_no_1_str"].isin(orders_with)) &
                (multi_df["product_name"] != sel_product)
            ]

            if paired_df.empty:
                st.info("함께 구매된 다른 상품이 없습니다.")
            else:
                paired_agg = paired_df.groupby("product_name").agg(
                    합포건수=("order_no_1_str", "nunique"),
                    판매수량=("net_qty", "sum"),
                    결제금액=("net_sales", "sum")
                ).reset_index().rename(columns={"product_name": "함께 구매한 상품"})
                paired_agg = paired_agg.sort_values("합포건수", ascending=False).reset_index(drop=True)
                paired_agg.index += 1

                st.markdown(f"<span style='color:{TEXT2};font-size:0.9rem;'>"
                            f"<b style='color:{CYAN};'>{sel_product}</b> 포함 합포 주문 "
                            f"<b style='color:{TEXT};'>{n_orders_with:,}건</b> 중 함께 구매한 상품 순위"
                            f"</span>", unsafe_allow_html=True)

                disp = paired_agg.copy()
                disp["결제금액"] = disp["결제금액"].apply(lambda x: f"₩{x:,}")
                disp.columns = ["함께 구매한 상품", "합포 건수", "판매수량", "결제금액"]
                st.dataframe(disp, use_container_width=True, hide_index=False, height=500)

# ════════════════════════════════════════════════════════
# PAGE 4 : 재고소진일정
# ════════════════════════════════════════════════════════
elif page == "📦 재고소진일정":

    st.markdown("# 📦 재고소진일정")
    st.markdown("---")

    today = date.today()

    # ── 주차 계산 (오늘 기준 이번주/다음주/다다음주) ─────
    monday_w1 = today - timedelta(days=today.weekday())
    monday_w2 = monday_w1 + timedelta(weeks=1)
    monday_w3 = monday_w1 + timedelta(weeks=2)
    sunday_w1 = monday_w1 + timedelta(days=6)
    sunday_w2 = monday_w2 + timedelta(days=6)
    sunday_w3 = monday_w3 + timedelta(days=6)
    w1_label  = f"{monday_w1.strftime('%m/%d')}~{sunday_w1.strftime('%m/%d')}"
    w2_label  = f"{monday_w2.strftime('%m/%d')}~{sunday_w2.strftime('%m/%d')}"
    w3_label  = f"{monday_w3.strftime('%m/%d')}~{sunday_w3.strftime('%m/%d')}"

    with st.expander("🔧 조회 조건", expanded=True):
        ec1, ec2, ec3 = st.columns([2, 2, 2])
        with ec1:
            preset = st.radio(
                "일평균 계산 기간",
                ["이번달", "지난달", "최근 30일", "최근 90일", "직접 입력"],
                horizontal=True, key="inv_preset"
            )
        with ec2:
            inv_file = st.file_uploader("📦 재고파일 업로드 (.xlsx)", type=["xlsx"], key="inv_upload")
            if inv_file is not None:
                st.session_state["inv_bytes"] = inv_file.read()
                st.session_state["inv_name"]  = inv_file.name
                st.session_state["inv_time"]  = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")
            if "inv_time" in st.session_state:
                st.caption(f"📁 {st.session_state['inv_name']}  |  {st.session_state['inv_time']}")
        with ec3:
            po_file = st.file_uploader("🚚 입고예정파일 업로드 (.xlsx)", type=["xlsx"], key="po_upload")
            if po_file is not None:
                st.session_state["po_bytes"] = po_file.read()
                st.session_state["po_name"]  = po_file.name
                st.session_state["po_time"]  = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")
            if "po_time" in st.session_state:
                st.caption(f"📁 {st.session_state['po_name']}  |  {st.session_state['po_time']}")

        # 기간 계산
        if preset == "이번달":
            period_from = today.replace(day=1)
            period_to   = today
        elif preset == "지난달":
            first_this  = today.replace(day=1)
            period_to   = first_this - timedelta(days=1)
            period_from = period_to.replace(day=1)
        elif preset == "최근 30일":
            period_from = today - timedelta(days=30)
            period_to   = today
        elif preset == "최근 90일":
            period_from = today - timedelta(days=90)
            period_to   = today
        else:
            dc1, dc2 = st.columns(2)
            with dc1:
                period_from = st.date_input("시작일", value=today - timedelta(days=30),
                                            min_value=df_sales["date"].min().date(),
                                            max_value=today, key="inv_from")
            with dc2:
                period_to = st.date_input("종료일", value=today,
                                          min_value=df_sales["date"].min().date(),
                                          max_value=today, key="inv_to")

        st.caption(f"📅 일평균 계산 기간: {period_from} ~ {period_to}  ({(period_to - period_from).days + 1}일)")

    # ── 일평균결제수량 계산 ───────────────────────────────
    days_avg = (period_to - period_from).days + 1
    sales_period = df_sales[
        (df_sales["date"].dt.date >= period_from) &
        (df_sales["date"].dt.date <= period_to)
    ]
    avg_by_pid = sales_period.groupby("product_id").agg(
        판매수량합=("net_qty", "sum")
    ).reset_index()
    avg_by_pid["product_id"] = avg_by_pid["product_id"].astype(str).str.strip()
    avg_by_pid["일평균결제수량"] = (avg_by_pid["판매수량합"] / days_avg).round(2)

    # ── 입고예정 데이터 처리 ─────────────────────────────
    po_bytes = st.session_state.get("po_bytes", None)
    incoming_by_pid = {}   # {product_id: [(date, qty), ...]}

    if po_bytes is not None:
        po_wb   = openpyxl.load_workbook(io.BytesIO(po_bytes), read_only=True, data_only=True)
        ws_po   = po_wb.active
        po_rows = list(ws_po.iter_rows(values_only=True))
        po_data = [r[:7] for r in po_rows[1:]
                   if r[0] is not None and str(r[0]).strip() not in ("", "품명", "상품명")]
        po_df   = pd.DataFrame(po_data, columns=["상품명","product_id","대분류","중분류","소분류","수량","예정일"])
        po_df["product_id"] = po_df["product_id"].astype(str).str.strip()
        po_df["수량"]   = pd.to_numeric(po_df["수량"], errors="coerce").fillna(0).astype(int)
        po_df["예정일"] = pd.to_datetime(po_df["예정일"], errors="coerce").dt.date
        po_df = po_df.dropna(subset=["예정일"])
        for pid, grp in po_df.groupby("product_id"):
            incoming_by_pid[str(pid)] = list(zip(grp["예정일"], grp["수량"]))

    def week_qty(pid, w_from, w_to):
        return sum(qty for d, qty in incoming_by_pid.get(str(pid), []) if w_from <= d <= w_to)

    # ── 소진일 타임라인 시뮬레이션 ──────────────────────
    def calc_depletion_timeline(stock, daily_rate, incoming_list):
        """입고예정을 반영한 타임라인 기반 소진일 계산"""
        if pd.isna(daily_rate) or daily_rate <= 0:
            return None
        remaining    = float(stock)
        current_date = today
        total_days   = 0
        for event_date, event_qty in sorted(incoming_list, key=lambda x: x[0]):
            if event_date <= current_date:
                remaining += event_qty   # 이미 지난 입고는 즉시 반영
                continue
            days_gap   = (event_date - current_date).days
            stock_used = daily_rate * days_gap
            if remaining <= stock_used:
                return total_days + int(remaining / daily_rate)
            remaining   -= stock_used
            remaining   += event_qty
            total_days  += days_gap
            current_date = event_date
        if remaining <= 0:
            return total_days
        return total_days + int(remaining / daily_rate)

    # ── 재고파일 처리 ─────────────────────────────────────
    inv_bytes = st.session_state.get("inv_bytes", None)

    if inv_bytes is None:
        st.info("👆 재고파일(.xlsx)을 업로드하면 소진 일정이 표시됩니다.")
        st.markdown(f"""
        <div style='background:{CARD};border:1px dashed {BORDER};border-radius:8px;padding:20px;margin-top:12px;'>
            <p style='color:{TEXT2};margin:0;font-size:0.9rem;'>
            📦 재고파일: <span style='color:{CYAN};'>상품명 / 품번 / 대분류 / 중분류 / 소분류 / 재고수량</span><br><br>
            🚚 입고예정파일: <span style='color:{CYAN};'>상품명 / 품번 / 대분류 / 중분류 / 소분류 / 수량 / 예정일</span>
            </p>
        </div>""", unsafe_allow_html=True)
    else:
        inv_wb   = openpyxl.load_workbook(io.BytesIO(inv_bytes), read_only=True, data_only=True)
        ws_inv   = inv_wb.active
        inv_rows = list(ws_inv.iter_rows(values_only=True))
        inv_data = [r[:6] for r in inv_rows[1:]
                    if r[0] is not None and str(r[0]).strip() not in ("", "품명", "상품명", "합계")]
        inv_df = pd.DataFrame(inv_data, columns=["상품명","product_id","대분류","중분류","소분류","재고수량"])
        inv_df["product_id"] = inv_df["product_id"].astype(str).str.strip()
        inv_df["재고수량"]   = pd.to_numeric(inv_df["재고수량"], errors="coerce").fillna(0).astype(int)

        result = inv_df.merge(avg_by_pid[["product_id","판매수량합","일평균결제수량"]], on="product_id", how="left")
        result["판매수량합"] = result["판매수량합"].fillna(0).astype(int)

        # 주차별 입고예정 수량 컬럼
        result["w1"] = result["product_id"].apply(lambda p: week_qty(p, monday_w1, sunday_w1))
        result["w2"] = result["product_id"].apply(lambda p: week_qty(p, monday_w2, sunday_w2))
        result["w3"] = result["product_id"].apply(lambda p: week_qty(p, monday_w3, sunday_w3))

        # 소진일 계산 (입고예정 타임라인 반영)
        result["잔여일수"] = result.apply(
            lambda r: calc_depletion_timeline(
                r["재고수량"],
                r["일평균결제수량"],
                incoming_by_pid.get(str(r["product_id"]), [])
            ), axis=1
        )
        result["소진예정일"] = result["잔여일수"].apply(
            lambda d: (today + timedelta(days=int(d))).strftime("%Y-%m-%d") if pd.notna(d) else "-"
        )
        result["일평균결제수량_표시"] = result["일평균결제수량"].apply(
            lambda v: f"{v:.1f}개/일" if pd.notna(v) else "결제수량 없음"
        )
        result["잔여일수_표시"] = result["잔여일수"].apply(
            lambda d: f"{int(d):,}일" if pd.notna(d) else "-"
        )

        def get_status(row):
            if pd.isna(row["잔여일수"]):
                return "⬜ 결제수량 없음"
            d = row["잔여일수"]
            if d <= 7:   return "🔴 긴급"
            elif d <= 30: return "🟡 주의"
            else:         return "🟢 여유"

        result["상태"] = result.apply(get_status, axis=1)

        status_order = {"🔴 긴급": 0, "🟡 주의": 1, "🟢 여유": 2, "⬜ 결제수량 없음": 3}
        result["_sort"] = result["상태"].map(status_order)
        result = result.sort_values(["_sort", "잔여일수"]).drop(columns="_sort").reset_index(drop=True)
        result.index += 1

        # ── 요약 스코어카드 ───────────────────────────────
        n_urgent  = (result["상태"] == "🔴 긴급").sum()
        n_caution = (result["상태"] == "🟡 주의").sum()
        n_safe    = (result["상태"] == "🟢 여유").sum()
        n_nodata  = (result["상태"] == "⬜ 결제수량 없음").sum()

        k1,k2,k3,k4 = st.columns(4)
        with k1: st.metric("🔴 긴급 (7일 이내)",   f"{n_urgent:,}개")
        with k2: st.metric("🟡 주의 (30일 이내)",   f"{n_caution:,}개")
        with k3: st.metric("🟢 여유 (30일 초과)",   f"{n_safe:,}개")
        with k4: st.metric("⬜ 결제수량 없음",       f"{n_nodata:,}개")

        st.markdown("---")

        # ── 결과 테이블 ───────────────────────────────────
        po_note = "✅ 입고예정 반영됨" if po_bytes else "⚠️ 입고예정파일 미업로드 (현재 재고 기준)"
        st.markdown(f"### 재고 소진 예정 현황  <span style='color:{TEXT2};font-size:0.85rem;'>전체 {len(result):,}개 품목 · {po_note}</span>", unsafe_allow_html=True)

        disp = result[[
            "상품명","product_id","대분류","중분류","소분류",
            "재고수량","판매수량합","일평균결제수량_표시",
            "w1","w2","w3",
            "잔여일수_표시","소진예정일","상태"
        ]].copy()
        disp.columns = [
            "상품명","품번","대분류","중분류","소분류",
            "재고수량", f"기간총출고량({days_avg}일)", "일평균결제수량",
            f"입고({w1_label})", f"입고({w2_label})", f"입고({w3_label})",
            "잔여일수","소진예정일","상태"
        ]
        st.dataframe(disp, use_container_width=True, hide_index=False, height=600)

        # ── 긴급 품목 별도 강조 ───────────────────────────
        urgent_df = result[result["상태"] == "🔴 긴급"]
        if not urgent_df.empty:
            st.markdown("---")
            st.markdown(f"### 🔴 긴급 소진 임박 품목 ({n_urgent}개)")
            urg_disp = urgent_df[[
                "상품명","product_id","재고수량","판매수량합","일평균결제수량_표시",
                "w1","w2","w3","잔여일수_표시","소진예정일"
            ]].copy()
            urg_disp.columns = [
                "상품명","품번","재고수량", f"기간총출고량({days_avg}일)", "일평균결제수량",
                f"입고({w1_label})", f"입고({w2_label})", f"입고({w3_label})",
                "잔여일수","소진예정일"
            ]
            urg_disp = urg_disp.reset_index(drop=True)
            urg_disp.index += 1
            st.dataframe(urg_disp, use_container_width=True, hide_index=False)

# ════════════════════════════════════════════════════════
# PAGE 5 : 챗봇
# ════════════════════════════════════════════════════════
elif page == "💬 챗봇":

    st.markdown("# 💬 데이터 챗봇")
    st.markdown("---")

    # 예시 질문 힌트
    st.markdown(f"""
    <div style='background:{CARD};border:1px solid {BORDER};border-radius:8px;padding:16px 20px;margin-bottom:20px;'>
        <p style='color:{TEXT2};font-size:0.85rem;margin:0 0 8px 0;'>💡 <b style='color:{TEXT};'>이런 질문을 해보세요</b></p>
        <div style='display:flex;flex-wrap:wrap;gap:8px;'>
            {"".join([f"<span style='background:#0e3a50;color:{CYAN};padding:4px 12px;border-radius:20px;font-size:0.82rem;'>{ex}</span>"
                      for ex in ["6월 21일 매출 얼마야?","이번달 판매수량 알려줘","지난달 채널별 결제금액","최근 30일 주문건수",
                                 "5월 상품 순위 TOP 5","오늘 매출","이번달 채월별 매출"]])}
        </div>
    </div>""", unsafe_allow_html=True)

    # 채팅 기록 초기화
    if "chat_history" not in st.session_state:
        st.session_state.chat_history = []

    # 대화 표시
    for msg in st.session_state.chat_history:
        with st.chat_message(msg["role"]):
            st.markdown(msg["content"])

    # 입력창
    if prompt := st.chat_input("질문을 입력하세요. 예: 6월 매출 얼마야?"):
        st.session_state.chat_history.append({"role": "user", "content": prompt})
        with st.chat_message("user"):
            st.markdown(prompt)

        answer = answer_question(prompt, df_sales, date.today())
        st.session_state.chat_history.append({"role": "assistant", "content": answer})
        with st.chat_message("assistant"):
            st.markdown(answer)

    # 초기화 버튼
    if st.session_state.chat_history:
        st.markdown("")
        if st.button("🗑️ 대화 초기화"):
            st.session_state.chat_history = []
            st.rerun()

st.markdown("---")
st.caption("Dotus Dashboard · 데이터 기준: 제공된 엑셀 파일")
